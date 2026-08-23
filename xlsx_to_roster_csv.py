#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""One-time helper: convert the official B.Tech 1st Year Session 2026-27
Excel roll list into the flat admission_data.csv consumed by
student_roster_import.py.

Source workbook layout (verified):
  * one sheet per section: CE A/B, CSE A-D, EE A/B, ECE A/B/C, IOT A,
    ME A/B, CHE A, IT A/B
  * header row contains: S.No | FORM NUMBER | Roll No. | Enrollment No |
    NAME OF THE APPLICANT | NAME OF FATHER | Section | T/P Batch
  * Roll No. arrives as a float (2026011001.0)

Output columns (same names student_roster_import.py expects, plus
Section/Batch so the app can place students correctly):
  S.No,Form_Number,Roll_No,Enrollment_No,Applicant_Name,Section,Batch,Source
"""
import csv
import re

import openpyxl

XLSX_PATH = r"d:\B.Tech 1st Year Session 2026-27.xlsx"
CSV_PATH = r"e:\mmmut\admission_data.csv"

ROLL_RE = re.compile(r"^\d{10}$")
ENROLL_RE = re.compile(r"^2026[A-Z]{3}\d{4}$")

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
rows_out = []
problems = []

for sheet in wb.sheetnames:
    ws = wb[sheet]
    header_row_idx = None
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        if row and str(row[0]).strip().rstrip(".").upper() == "S.NO":
            header_row_idx = i
            break
    if header_row_idx is None:
        problems.append(f"{sheet}: no S.No header row found")
        continue
    for r in rows[header_row_idx + 1:]:
        if not r or r[0] is None:
            continue
        sno = str(r[0]).strip()
        form = str(r[1] or "").strip()
        roll_raw = r[2]
        roll = str(int(roll_raw)) if isinstance(roll_raw, float) else str(roll_raw or "").strip()
        enroll = str(r[3] or "").strip()
        name = re.sub(r"\s+", " ", str(r[4] or "")).strip()
        section = str(r[6] or "").strip()
        batch = str(r[7] or "").strip()
        if not ROLL_RE.match(roll):
            problems.append(f"{sheet}: bad roll '{roll}' ({name})")
            continue
        if not ENROLL_RE.match(enroll):
            problems.append(f"{sheet}: bad enrollment '{enroll}' ({name})")
            continue
        if not name:
            problems.append(f"{sheet}: empty name for roll {roll}")
            continue
        rows_out.append({
            "S.No": sno,
            "Form_Number": form,
            "Roll_No": roll,
            "Enrollment_No": enroll,
            "Applicant_Name": name.upper(),
            "Section": section,
            "Batch": batch,
            "Source": sheet,
        })

fieldnames = ["S.No", "Form_Number", "Roll_No", "Enrollment_No",
              "Applicant_Name", "Section", "Batch", "Source"]
with open(CSV_PATH, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows_out)

seen = {}
for rec in rows_out:
    seen.setdefault(rec["Roll_No"], []).append(rec["Enrollment_No"])
dupes = {k: v for k, v in seen.items() if len(v) > 1}

by_prefix = {}
for rec in rows_out:
    prefix = rec["Enrollment_No"][4:7]
    by_prefix[prefix] = by_prefix.get(prefix, 0) + 1

print(f"Sheets read     : {len(wb.sheetnames)}")
print(f"Students written: {len(rows_out)} -> {CSV_PATH}")
print("Per-branch tally:", dict(sorted(by_prefix.items())))
print("Duplicate rolls :", len(dupes))
for p in problems[:10]:
    print("  problem:", p)
print("Problems total  :", len(problems))
